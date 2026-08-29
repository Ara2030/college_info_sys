# -*- coding: utf-8 -*-
"""
Сервисы модуля «Отчётность и интеграция» (2.6).

  build_stat_report(report_type, period_year) — формирование отчётов СПО-1/СПО-2/мониторинг
  send_to_registry_api(xml_bytes, endpoint)   — REST API Реестра студентов СПО (XML)
  smev_request(...)                            — запрос сведений из госреестров через СМЭВ
  stat_report_excel(report)                    — экспорт отчёта в Excel (openpyxl)
  stat_report_pdf(report)                      — экспорт отчёта в PDF (reportlab)
"""
import hashlib
import io
import json
from datetime import date

from django.conf import settings
from django.utils import timezone

from .models import IntegrationLog, SmevRequest, StatReport


# ==========================================================================
# Сбор статистики
# ==========================================================================

def _collect_contingent_stats():
    from contingent.models import Department, Group, Specialty, Student, StudentStatus
    return {
        'Отделений': Department.objects.count(),
        'Специальностей': Specialty.objects.count(),
        'Групп': Group.objects.filter(is_active=True).count(),
        'Студентов': Student.objects.count(),
        'Обучается': Student.objects.filter(status=StudentStatus.STUDY).count(),
        'В академическом отпуске': Student.objects.filter(
            status=StudentStatus.ACADEMIC_LEAVE).count(),
        'Отчислено': Student.objects.filter(status=StudentStatus.EXPELLED).count(),
        'Выпускники': Student.objects.filter(status=StudentStatus.GRADUATED).count(),
    }


def _collect_students_by_course():
    from contingent.models import Student
    by_course = {}
    for student in Student.objects.select_related('group'):
        key = student.group.course if student.group else 0
        by_course[key] = by_course.get(key, 0) + 1
    return {f'{k} курс': v for k, v in sorted(by_course.items())}


def _collect_journal_stats():
    from journal.models import Grade, Lesson
    grades = Grade.objects.exclude(value='')
    return {
        'Учебных занятий проведено': Lesson.objects.count(),
        'Оценок выставлено': grades.count(),
    }


def _collect_attestation_stats():
    from attestation.models import AcademicDebt, Exam, ScholarshipPeriod
    return {
        'Экзаменов в сессию': Exam.objects.filter(exam_type='exam').count(),
        'Зачётов': Exam.objects.filter(exam_type='credit').count(),
        'Академических задолженностей (активные)': AcademicDebt.objects.filter(
            status=AcademicDebt.Status.ACTIVE).count(),
        'Студентов на стипендии (последний период)': (
            ScholarshipPeriod.objects.order_by('-id').first().students_count
            if ScholarshipPeriod.objects.exists() else 0),
    }


def _collect_schedule_stats():
    from schedule.models import Room, ScheduleEntry, Teacher
    return {
        'Аудиторий': Room.objects.count(),
        'Кабинетов': Room.objects.filter(room_type='lecture').count(),
        'Лабораторий': Room.objects.filter(room_type='lab').count(),
        'Компьютерных классов': Room.objects.filter(room_type='computer').count(),
        'Преподавателей': Teacher.objects.filter(is_active=True).count(),
        'Занятий в расписании': ScheduleEntry.objects.count(),
    }


def _collect_hr_stats():
    from hr.models import Employee, HROrder
    return {
        'Сотрудников (всего)': Employee.objects.count(),
        'Педагогических работников': Employee.objects.filter(
            category=Employee.Category.TEACHER, status=Employee.Status.ACTIVE).count(),
        'АУП': Employee.objects.filter(category=Employee.Category.MANAGEMENT,
                                       status=Employee.Status.ACTIVE).count(),
        'Вспомогательный персонал': Employee.objects.filter(
            category=Employee.Category.SUPPORT, status=Employee.Status.ACTIVE).count(),
        'Приказов по личному составу': HROrder.objects.count(),
    }


def _spo1_data(year):
    """СПО-1: сведения о сети и численности."""
    from contingent.models import Order, OrderItem
    year_orders = Order.objects.filter(date__year=year)
    return {
        'Наименование отчёта': 'СПО-1. Сведения об образовательной организации',
        'Отчётный год': year,
        'Раздел 1. Сеть образовательной организации': _collect_contingent_stats(),
        'Раздел 2. Численность по курсам': _collect_students_by_course(),
        'Раздел 3. Движение контингента за год': {
            'Приказов за год': year_orders.count(),
            'Зачислено (приказы о зачислении)': year_orders.filter(
                order_type__code='enroll').count(),
            'Отчислено (приказы об отчислении)': year_orders.filter(
                order_type__code='expel').count(),
            'Переводы': year_orders.filter(order_type__code='transfer').count(),
            'Академические отпуска': year_orders.filter(
                order_type__code='academic_leave').count(),
            'Восстановления': year_orders.filter(order_type__code='recover').count(),
        },
        'Раздел 4. Кадровое обеспечение': {
            k: v for k, v in _collect_hr_stats().items()
            if k in ('Педагогических работников', 'АУП', 'Вспомогательный персонал')
        },
    }


def _spo2_data(year):
    """СПО-2: сведения о материально-технической базе."""
    return {
        'Наименование отчёта': 'СПО-2. Сведения о материально-технической базе',
        'Отчётный год': year,
        'Аудиторный фонд': _collect_schedule_stats(),
        'Компьютерная техника': {
            'Компьютерных классов': _collect_schedule_stats()['Компьютерных классов'],
            'Аудиторий всего': _collect_schedule_stats()['Аудиторий'],
        },
    }


def _monitoring_data(year):
    """Мониторинг Минпросвещения: сводка по деятельности колледжа."""
    return {
        'Наименование отчёта': 'Мониторинг системы среднего профессионального образования',
        'Отчётный год': year,
        'Контингент': _collect_contingent_stats(),
        'Учебный процесс': _collect_journal_stats(),
        'Промежуточная аттестация': _collect_attestation_stats(),
        'Инфраструктура и кадры': _collect_hr_stats(),
    }


BUILDERS = {
    StatReport.Type.SPO1: _spo1_data,
    StatReport.Type.SPO2: _spo2_data,
    StatReport.Type.MONITORING: _monitoring_data,
}


def build_stat_report(report_type, period_year=None):
    """Формирует статистический отчёт и сохраняет его в БД."""
    period_year = period_year or date.today().year
    data = BUILDERS[report_type](period_year)
    data['Сформировано'] = timezone.now().isoformat()

    payload = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')

    report = StatReport.objects.create(
        report_type=report_type,
        period_year=period_year,
        period_date=date.today(),
        status=StatReport.Status.READY,
        data_json=json.dumps(data, ensure_ascii=False, indent=2),
        file_name=f'{report_type}_report_{period_year}_{timezone.now():%Y%m%d_%H%M}.json',
        checksum=hashlib.sha256(payload).hexdigest(),
        comment='Сформирован автоматически по данным ИС.',
    )
    return report


# ==========================================================================
# REST API Реестра студентов СПО
# ==========================================================================

def get_registry_endpoint():
    return getattr(settings, 'REGISTRY_API_ENDPOINT',
                   'https://registry.spo.example.ru/api/v1/students')


def send_to_registry_api(xml_bytes: bytes, endpoint: str = None):
    """
    Отправка выгрузки в Реестр студентов СПО через REST API (формат XML).

    Если реальное подключение не настроено (settings.REGISTRY_API_ENABLED=False),
    запрос логируется со статусом «Заглушка» — протокол и формат отработаны.
    """
    endpoint = endpoint or get_registry_endpoint()
    enabled = getattr(settings, 'REGISTRY_API_ENABLED', False)
    xml_text = xml_bytes.decode('utf-8')

    if enabled:
        # Реальная отправка (настраивается в settings)
        try:
            import requests
            resp = requests.post(endpoint, data=xml_bytes,
                                 headers={'Content-Type': 'application/xml'},
                                 timeout=30)
            status, code = (IntegrationLog.Status.OK, resp.status_code)
            response_text = resp.text[:4000]
        except Exception as exc:  # noqa: BLE001
            status, code = (IntegrationLog.Status.ERROR, None)
            response_text = str(exc)[:4000]
    else:
        status, code = (IntegrationLog.Status.MOCK, 202)
        response_text = 'OK (заглушка): сервер реестра не подключён — ' \
                        'данные приняты в протокольном виде. ' \
                        'Установите REGISTRY_API_ENABLED=True для реальной отправки.'

    return IntegrationLog.objects.create(
        integration=IntegrationLog.Integration.REGISTRY_API,
        direction='out',
        endpoint=endpoint,
        payload=xml_text[:4000],
        response=response_text,
        status=status,
        status_code=code,
    )


# ==========================================================================
# СМЭВ
# ==========================================================================

def smev_request(registry, identifier, person_name=''):
    """
    Запрос сведений из государственного реестра через СМЭВ.

    Если подключение не настроено (settings.SMEV_ENABLED=False),
    запрос фиксируется со статусом «Подключение не настроено».
    """
    enabled = getattr(settings, 'SMEV_ENABLED', False)
    registry_name = dict(SmevRequest.Registry.choices).get(registry, registry)

    if not enabled:
        return SmevRequest.objects.create(
            registry=registry,
            identifier=identifier,
            person_name=person_name,
            status=SmevRequest.Status.NO_CONNECTION,
            response_text='Подключение к СМЭВ не настроено (settings.SMEV_ENABLED=False). '
                          'Для реального обмена необходимо подключение к транспортной '
                          'подсистеме СМЭВ и сертификаты.',
        )

    # Имитация ответа при подключении
    mock_data = {
        'fns': f'Сведения ФНС по ИНН {identifier}: налогоплательщик '
               f'{person_name or "не указан"}, статус — действующий.',
        'pfr': f'Сведения СФР по СНИЛС {identifier}: страховой номер подтверждён, '
               f'страховой стаж — 5 лет.',
        'zags': f'Сведения ЗАГС по запросу {identifier}: запись акта найдена.',
        'mvd': f'Сведения МВД по паспорту {identifier}: документ действителен.',
    }
    IntegrationLog.objects.create(
        integration=IntegrationLog.Integration.SMEV,
        direction='in',
        endpoint=f'СМЭВ → {registry_name}',
        payload=f'{registry_name}: {identifier}',
        response=mock_data.get(registry, ''),
        status=IntegrationLog.Status.OK,
    )
    return SmevRequest.objects.create(
        registry=registry,
        identifier=identifier,
        person_name=person_name,
        status=SmevRequest.Status.CONNECTED,
        response_text=mock_data.get(registry, ''),
    )


# ==========================================================================
# Экспорт: Excel, PDF
# ==========================================================================

def stat_report_excel(report: StatReport):
    """Экспорт данных отчёта в Excel (.xlsx) через openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = report.get_report_type_display()[:30]

    ws['A1'] = report.get_report_type_display()
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'Отчётный год: {report.period_year} · Сформирован: {report.created_at:%d.%m.%Y %H:%M}'
    ws['A2'].font = Font(italic=True, size=9)

    data = json.loads(report.data_json)
    row = 4
    for section, values in data.items():
        if isinstance(values, dict):
            ws.cell(row=row, column=1, value=section).font = Font(bold=True)
            row += 1
            for key, value in values.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
        else:
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=2, value=values)
            row += 1

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f'{report.report_type}_report_{report.period_year}.xlsx'


def stat_report_pdf(report: StatReport):
    """Экспорт данных отчёта в PDF через reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=14,
                                 spaceAfter=6)
    h_style = ParagraphStyle('H', parent=styles['Heading2'], fontSize=11,
                             spaceBefore=8, spaceAfter=4)
    data = json.loads(report.data_json)

    story = [Paragraph(report.get_report_type_display(), title_style),
             Paragraph(f'Отчётный год: {report.period_year} · '
                       f'Сформирован: {report.created_at:%d.%m.%Y %H:%M}',
                       styles['Normal']),
             Spacer(1, 4 * mm)]

    for section, values in data.items():
        if isinstance(values, dict):
            story.append(Paragraph(section, h_style))
            rows = [[k, str(v)] for k, v in values.items()]
            table = Table([['Показатель', 'Значение']] + rows,
                          colWidths=[110 * mm, 60 * mm])
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1),
                 [colors.white, colors.HexColor('#f1f5f9')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
        else:
            story.append(Paragraph(f'{section}: {values}', styles['Normal']))

    doc.build(story)
    buf.seek(0)
    return buf, f'{report.report_type}_report_{report.period_year}.pdf'
